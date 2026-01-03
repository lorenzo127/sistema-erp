import datetime
import io
import json
import os
import pandas as pd
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# --- IMPORTS DJANGO ---
from django.conf import settings
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.models import Group
from django.contrib.auth.views import LoginView
from django.core.paginator import Paginator
from django.core.mail import send_mail
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.db.models import Avg, Count, IntegerField, Q, Sum
from django.db.models.functions import Cast, TruncDay, TruncMonth
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.utils import timezone

# --- IMPORTS TERCEROS ---
try:
    from weasyprint import HTML
except (ImportError, OSError):
    HTML = None

# --- IMPORTS LOCALES ---
from .models import (
    CajaChica, 
    CentroCosto, 
    Clasificacion, 
    Empresa, 
    Ingreso, 
    Trabajador, 
    Cargo, 
    Movimiento,
    Producto, 
    Lote      
)

from .forms import (
    CajaChicaForm,
    CargaExcelForm,
    IngresoForm,
    RegistroUsuarioForm,
    TrabajadorForm,
    LoteForm,
    SalidaStockForm
)

from .services import DashboardService
from .ia import entrenar_modelo, predecir_categoria

# =========================================================
# 0. FUNCIONES DE SEGURIDAD (Permisos)
# =========================================================
def es_finanzas(user):
    # Pasa si es Superusuario O pertenece al grupo Finanzas
    return user.is_superuser or user.groups.filter(name='Finanzas').exists()

def es_bodega(user):
    return user.is_superuser or user.groups.filter(name='Bodega').exists()

def es_rrhh(user):
    return user.is_superuser or user.groups.filter(name='RRHH').exists()


# =========================================================
# 1. DASHBOARD GENERAL (Página de Inicio)
# =========================================================
@login_required
def dashboard(request):
    """VISTA PRINCIPAL: COMANDO CENTRAL"""
    hoy = datetime.date.today()
    inicio_mes = hoy.replace(day=1)

    # Definimos qué palabras clave cuentan como dinero entrando
    tipos_entrada = ['INGRESO', 'VENTA', 'ABONO', 'DEVOLUCION']

    # 1. KPI FINANCIEROS (Mes Actual)
    # A. GASTOS: Sumamos todo lo que NO sea entrada
    total_gastos = Ingreso.objects.filter(
        fecha__gte=inicio_mes
    ).exclude(tipo_documento__in=tipos_entrada).aggregate(total=Sum('monto_transferencia'))['total'] or 0

    # B. INGRESOS: Sumamos solo lo que sea entrada
    total_ingresos = Ingreso.objects.filter(
        fecha__gte=inicio_mes,
        tipo_documento__in=tipos_entrada
    ).aggregate(total=Sum('monto_transferencia'))['total'] or 0

    resultado_mes = total_ingresos - total_gastos

    # 2. KPI INVENTARIO
    stock_vencido = Lote.objects.filter(fecha_vencimiento__lt=hoy).count()
    limite_alerta = hoy + datetime.timedelta(days=30)
    stock_critico = Lote.objects.filter(
        fecha_vencimiento__gte=hoy, 
        fecha_vencimiento__lte=limite_alerta
    ).count()

    # 3. KPI RRHH
    try:
        personal_activo = Trabajador.objects.filter(fecha_finiquito__isnull=True).count()
    except:
        personal_activo = 0

    context = {
        'total_gastos': total_gastos,
        'total_ingresos': total_ingresos,
        'resultado_mes': resultado_mes,
        'stock_vencido': stock_vencido,
        'stock_critico': stock_critico,
        'personal_activo': personal_activo,
        'fecha_actual': hoy,
    }
    return render(request, 'core/dashboard.html', context)

# =========================================================
# 2. MÓDULO FINANZAS (Control de Movimientos .xlsm)
# =========================================================
@login_required
@user_passes_test(es_finanzas)
def finanzas_dashboard(request):
    """Dashboard Financiero con Filtros de Fecha."""
    
    # 1. Capturar Filtros
    anio = request.GET.get('anio')
    mes = request.GET.get('mes')

    # Queryset base (todos los movimientos)
    queryset = Movimiento.objects.all()

    # Aplicar filtros si existen
    if anio:
        queryset = queryset.filter(fecha__year=anio)
    
    if anio and mes: 
        queryset = queryset.filter(fecha__month=mes)

    # 2. Calcular KPIs
    total_ingresos = queryset.filter(tipo='INGRESO').aggregate(Sum('monto'))['monto__sum'] or 0
    total_egresos = queryset.filter(tipo='EGRESO').aggregate(Sum('monto'))['monto__sum'] or 0
    balance = total_ingresos - total_egresos

    # 3. Datos para Gráfico de Evolución
    if anio and mes:
        # Agrupar por DÍA
        evolucion = queryset.annotate(fecha_trunc=TruncDay('fecha'))\
                            .values('fecha_trunc')\
                            .annotate(
                                ingreso=Sum('monto', filter=Q(tipo='INGRESO')),
                                egreso=Sum('monto', filter=Q(tipo='EGRESO'))
                            ).order_by('fecha_trunc')
        formato_fecha = "%d %b"
    else:
        # Agrupar por MES
        evolucion = queryset.annotate(fecha_trunc=TruncMonth('fecha'))\
                            .values('fecha_trunc')\
                            .annotate(
                                ingreso=Sum('monto', filter=Q(tipo='INGRESO')),
                                egreso=Sum('monto', filter=Q(tipo='EGRESO'))
                            ).order_by('fecha_trunc')
        formato_fecha = "%B %Y"

    labels_evolucion = []
    data_ingresos = []
    data_egresos = []

    for e in evolucion:
        if e['fecha_trunc']:
            labels_evolucion.append(e['fecha_trunc'].strftime(formato_fecha))
            data_ingresos.append(e['ingreso'] or 0)
            data_egresos.append(e['egreso'] or 0)

    # 4. Obtener Años Disponibles
    anios_disponibles = Movimiento.objects.dates('fecha', 'year', order='DESC')

    context = {
        'total_ingresos': total_ingresos,
        'total_egresos': total_egresos,
        'balance': balance,
        'movimientos': queryset.order_by('-fecha')[:50],
        'pie_labels': ['Ingresos', 'Egresos'],
        'pie_data': [total_ingresos, total_egresos],
        'bar_labels': labels_evolucion,
        'bar_ingresos': data_ingresos,
        'bar_egresos': data_egresos,
        'anios_disponibles': anios_disponibles,
        'anio_seleccionado': int(anio) if anio else None,
        'mes_seleccionado': int(mes) if mes else None,
    }
    return render(request, 'core/finanzas/dashboard.html', context)

@login_required
def importar_finanzas(request):
    """Importador Específico para Hoja 'Control de Finanzas'"""
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        archivo = request.FILES['archivo_excel']
        try:
            df = pd.read_excel(
                archivo, 
                engine='openpyxl', 
                sheet_name='Control de Finanzas',
                header=12,
                usecols="B:F"
            )
            
            nuevos_nombres = ['FECHA', 'DESCRIPCION', 'TIPO', 'CATEGORIA', 'MONTO']
            if len(df.columns) == 5:
                df.columns = nuevos_nombres
            
            creados = 0
            
            with transaction.atomic():
                for index, row in df.iterrows():
                    fecha = row.get('FECHA')
                    if pd.isnull(fecha) or str(fecha).strip() == '': continue

                    desc = str(row.get('DESCRIPCION', '')).strip()
                    if desc == 'nan': desc = 'Sin detalle'

                    categoria = str(row.get('CATEGORIA', '')).strip()
                    if categoria and categoria != 'nan':
                        desc = f"{categoria} - {desc}"

                    try:
                        val_monto = row.get('MONTO', 0)
                        if isinstance(val_monto, str):
                             val_monto = val_monto.replace('$', '').replace('.', '').replace(',', '')
                        monto = abs(int(float(val_monto)))
                    except:
                        monto = 0
                        
                    if monto == 0: continue

                    tipo_texto = str(row.get('TIPO', '')).upper()
                    tipo_final = 'EGRESO' 
                    if 'INGRESO' in tipo_texto or 'ABONO' in tipo_texto:
                        tipo_final = 'INGRESO'
                    
                    Movimiento.objects.create(
                        fecha=fecha,
                        descripcion=desc,
                        monto=monto,
                        tipo=tipo_final,
                    )
                    creados += 1

            if creados > 0:
                messages.success(request, f'¡Excelente! Se cargaron {creados} registros.')
            else:
                messages.warning(request, 'Se leyó la hoja, pero no se encontraron filas válidas.')
                
            return redirect('finanzas_dashboard')

        except Exception as e:
            if "Worksheet" in str(e) and "does not exist" in str(e):
                messages.error(request, 'Error: No se encontró la hoja llamada "Control de Finanzas".')
            else:
                messages.error(request, f"Error técnico: {str(e)}")
            print(f"Error Finanzas: {e}")

    return render(request, 'core/finanzas/importar.html')


# =========================================================
# 3. MÓDULO INGRESOS / GASTOS (CRUD Clásico)
# =========================================================
@login_required
@user_passes_test(es_finanzas)
def lista_ingresos(request):
    # 1. Base Query
    movimientos = Ingreso.objects.select_related('empresa', 'centro_costo', 'clasificacion').all()
    
    # 2. Filtros
    # Búsqueda texto
    q = request.GET.get('q')
    if q:
        movimientos = movimientos.filter(
            Q(descripcion_movimiento__icontains=q) |
            Q(detalle__icontains=q) |
            Q(empresa__nombre__icontains=q)
        )

    # Filtros Dropdown
    empresa_id = request.GET.get('empresa')
    if empresa_id: movimientos = movimientos.filter(empresa_id=empresa_id)

    centro_id = request.GET.get('centro')
    if centro_id: movimientos = movimientos.filter(centro_costo_id=centro_id)

    clasif_id = request.GET.get('clasificacion')
    if clasif_id: movimientos = movimientos.filter(clasificacion_id=clasif_id)

    # Filtros Fecha y Monto
    f_inicio = request.GET.get('fecha_inicio')
    f_fin = request.GET.get('fecha_fin')
    
    if f_inicio: movimientos = movimientos.filter(fecha__gte=f_inicio)
    if f_fin: movimientos = movimientos.filter(fecha__lte=f_fin)

    min_costo = request.GET.get('min_costo')
    max_costo = request.GET.get('max_costo')
    if min_costo: movimientos = movimientos.filter(monto_transferencia__gte=min_costo)
    if max_costo: movimientos = movimientos.filter(monto_transferencia__lte=max_costo)

    # 3. Ordenamiento
    orden = request.GET.get('orden', 'fecha_desc')
    if orden == 'fecha_asc': movimientos = movimientos.order_by('fecha')
    elif orden == 'fecha_desc': movimientos = movimientos.order_by('-fecha')
    elif orden == 'monto_asc': movimientos = movimientos.order_by('monto_transferencia')
    elif orden == 'monto_desc': movimientos = movimientos.order_by('-monto_transferencia')
    else: movimientos = movimientos.order_by('-fecha')

    # 4. Preparar Datos para el Gráfico (LÓGICA INTELIGENTE DÍA/MES)
    agrupar_por_dia = False
    
    if f_inicio and f_fin:
        try:
            d1 = datetime.datetime.strptime(f_inicio, '%Y-%m-%d')
            d2 = datetime.datetime.strptime(f_fin, '%Y-%m-%d')
            dias_diff = abs((d2 - d1).days)
            if dias_diff <= 60:
                agrupar_por_dia = True
        except ValueError:
            pass 

    if agrupar_por_dia:
        datos_grafico = movimientos.annotate(periodo=TruncDay('fecha'))\
                                   .values('periodo')\
                                   .annotate(total=Sum('monto_transferencia'))\
                                   .order_by('periodo')
        labels_grafico = [d['periodo'].strftime('%d/%m') for d in datos_grafico] if datos_grafico else []
    else:
        datos_grafico = movimientos.annotate(periodo=TruncMonth('fecha'))\
                                   .values('periodo')\
                                   .annotate(total=Sum('monto_transferencia'))\
                                   .order_by('periodo')
        labels_grafico = [d['periodo'].strftime('%Y-%m') for d in datos_grafico] if datos_grafico else []

    data_grafico = [d['total'] for d in datos_grafico] if datos_grafico else []

    # 5. Paginación
    per_page = request.GET.get('per_page', 25)
    paginator = Paginator(movimientos, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # --- RESPUESTA AJAX ---
    if request.GET.get('modo_ajax'):
        html_tabla = render_to_string('core/partials/tabla_ingresos.html', {'page_obj': page_obj}, request=request)
        html_paginacion = render_to_string('core/partials/paginacion.html', {'page_obj': page_obj}, request=request)
        
        return JsonResponse({
            'html_tabla': html_tabla,
            'html_paginacion': html_paginacion,
            'grafico_labels': labels_grafico,
            'grafico_data': data_grafico
        })

    # --- RESPUESTA NORMAL ---
    context = {
        'page_obj': page_obj,
        'empresas': Empresa.objects.all(),
        'centros': CentroCosto.objects.all(),
        'clasificaciones': Clasificacion.objects.all(),
        'labels_grafico': labels_grafico,
        'data_grafico': data_grafico,
        'orden_sel': orden,
        'per_page': int(per_page),
        'inicio_sel': f_inicio,
        'fin_sel': f_fin,
    }
    return render(request, 'core/lista_ingresos.html', context)

@login_required
def editar_ingreso(request, id):
    ingreso = get_object_or_404(Ingreso, id=id)
    if request.method == 'POST':
        form = IngresoForm(request.POST, instance=ingreso)
        if form.is_valid():
            form.save()
            messages.success(request, 'Registro actualizado correctamente.')
            return redirect('lista_ingresos')
    else:
        form = IngresoForm(instance=ingreso)
    return render(request, 'core/editar_ingreso.html', {'form': form, 'ingreso': ingreso})

@login_required
def eliminar_ingreso(request, id):
    ingreso = get_object_or_404(Ingreso, id=id)
    ingreso.delete()
    messages.success(request, 'Registro eliminado correctamente.')
    return redirect('lista_ingresos')

@login_required
def importar_excel(request):
    """Importador con AUTO-CREACIÓN de Categorías y Centros de Costo"""
    if request.method == 'POST':
        form = CargaExcelForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo_excel']
            creados = 0
            
            try:
                try:
                    df = pd.read_excel(archivo, sheet_name='REGISTRO EGRESOS', header=5)
                except:
                    df = pd.read_excel(archivo, header=5)

                df.columns = df.columns.str.strip()

                if 'Fecha' not in df.columns or 'Monto Transferencia' not in df.columns:
                    messages.error(request, 'Error: No se encontraron columnas "Fecha" o "Monto Transferencia" en la fila 6.')
                    return redirect('importar_excel')

                with transaction.atomic():
                    for index, row in df.iterrows():
                        fecha = row.get('Fecha')
                        if pd.isnull(fecha): continue
                        
                        monto = row.get('Monto Transferencia', 0)
                        if pd.isnull(monto) or monto == 0: continue

                        # 1. Empresa
                        nombre_empresa = str(row.get('Empresa', '')).strip()
                        empresa_obj = None
                        if nombre_empresa and nombre_empresa.lower() != 'nan':
                            empresa_obj, created = Empresa.objects.get_or_create(
                                nombre__iexact=nombre_empresa, 
                                defaults={'nombre': nombre_empresa}
                            )

                        # 2. Centro de Costo
                        nombre_centro = str(row.get('Centro de Costo', '')).strip()
                        centro_obj = None
                        if nombre_centro and nombre_centro.lower() != 'nan':
                            centro_obj, _ = CentroCosto.objects.get_or_create(
                                nombre__iexact=nombre_centro,
                                defaults={'nombre': nombre_centro}
                            )

                        # 3. Clasificación
                        nombre_clasif = str(row.get('Clasificación', '')).strip()
                        clasif_obj = None
                        if nombre_clasif and nombre_clasif.lower() != 'nan':
                            clasif_obj, _ = Clasificacion.objects.get_or_create(
                                nombre__iexact=nombre_clasif,
                                defaults={'nombre': nombre_clasif}
                            )

                        desc_movimiento = str(row.get('Descripcion de Movimiento', 'Sin descripción')).strip()
                        if desc_movimiento.lower() == 'nan': desc_movimiento = 'Sin descripción'

                        detalle_txt = str(row.get('Detalle', '')).strip()
                        if detalle_txt.lower() == 'nan': detalle_txt = ''
                        
                        n_doc = row.get('N° DOCUMENTO')
                        if not pd.isnull(n_doc):
                            detalle_txt = f"Doc: {n_doc} - {detalle_txt}"

                        tipo_doc = str(row.get('Tipo', 'GASTO')).strip()
                        
                        Ingreso.objects.create(
                            fecha=fecha,
                            monto_transferencia=monto,
                            descripcion_movimiento=desc_movimiento,
                            tipo_documento=tipo_doc,
                            detalle=detalle_txt,
                            empresa=empresa_obj,
                            centro_costo=centro_obj,
                            clasificacion=clasif_obj,
                            iva=0
                        )
                        creados += 1

                messages.success(request, f'¡Listo! Se cargaron {creados} registros y se crearon las categorías faltantes automáticamente.')

            except Exception as e:
                messages.error(request, f"Error técnico: {str(e)}")
                print(f"Error completo: {e}")
                
            return redirect('importar_excel')
    else:
        form = CargaExcelForm()
        
    return render(request, 'core/importar.html', {'form': form})

@login_required
def descargar_plantilla(request):
    ejemplo = {
        'Fecha': ['01/12/2025'],
        'Empresa': ['Nombre Empresa'],
        'Centro de Costo': ['Administracion'],
        'Clasificación': ['Insumos'],
        'N° DOCUMENTO': ['12345'],
        'Monto Transferencia': [50000],
        'Descripcion de Movimiento': ['Compra'],
        'Estado': ['Pagado'],
        'Detalle': ['Papeleria'],
        'Tipo': ['GASTO']
    }
    df = pd.DataFrame(ejemplo)
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='REGISTRO EGRESOS', index=False, startrow=5)
    
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=plantilla_importacion.xlsx'
    return response


# =========================================================
# 4. MÓDULO CAJA CHICA
# =========================================================
@login_required
@user_passes_test(es_finanzas)
def lista_caja_chica(request):
    gastos = CajaChica.objects.all().order_by('-fecha')

    resumen_meses = CajaChica.objects.annotate(mes=TruncMonth('fecha'))\
                                     .values('mes')\
                                     .annotate(total=Sum('monto'))\
                                     .order_by('mes')

    labels = []
    data = []
    
    for registro in resumen_meses:
        if registro['mes']:
            labels.append(registro['mes'].strftime('%Y-%m')) 
            monto_entero = int(registro['total']) 
            data.append(monto_entero)

    context = {
        'gastos': gastos,
        'labels_grafico': labels,
        'data_grafico': data,
    }
    return render(request, 'core/caja_chica_lista.html', context)

@login_required
def caja_chica_crear(request):
    if request.method == 'POST':
        form = CajaChicaForm(request.POST, request.FILES)
        if form.is_valid():
            gasto = form.save(commit=False)
            gasto.responsable = request.user
            gasto.save()
            messages.success(request, 'Gasto registrado correctamente.')
            return redirect('lista_caja_chica')
    else:
        form = CajaChicaForm()
    return render(request, 'core/caja_chica_form.html', {'form': form, 'titulo': 'Nuevo Gasto'})

@login_required
def caja_chica_editar(request, id):
    gasto = get_object_or_404(CajaChica, id=id)
    if request.method == 'POST':
        form = CajaChicaForm(request.POST, request.FILES, instance=gasto)
        if form.is_valid():
            form.save()
            messages.success(request, 'Gasto actualizado.')
            return redirect('lista_caja_chica')
    else:
        form = CajaChicaForm(instance=gasto)
    return render(request, 'core/caja_chica_form.html', {'form': form, 'titulo': 'Editar Gasto'})

@login_required
def caja_chica_eliminar(request, id):
    gasto = get_object_or_404(CajaChica, id=id)
    gasto.delete()
    messages.success(request, 'Gasto eliminado.')
    return redirect('lista_caja_chica')

@login_required
def exportar_caja_chica_pdf(request):
    gastos = CajaChica.objects.all().order_by('-fecha')
    total_gasto = gastos.aggregate(Sum('monto'))['monto__sum'] or 0
    
    context = {
        'gastos': gastos,
        'total_gasto': total_gasto,
        'fecha_emision': datetime.datetime.now(),
        'usuario': request.user,
        'empresa_nombre': "SAMKA / MAQUEHUE",
    }
    
    html_string = render_to_string('core/pdf/reporte_caja.html', context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="rendicion_caja_chica.pdf"'
    
    if HTML:
        HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf(response)
    else:
        return HttpResponse("Error: La librería de PDF no está configurada (Falta GTK).")
    return response


# =========================================================
# 5. MÓDULO RRHH (Trabajadores y Finiquitos)
# =========================================================
@login_required
@user_passes_test(es_rrhh)
def dashboard_rrhh(request):
    filtro_empresa = request.GET.get('empresa', '')
    workers_queryset = Trabajador.objects.all()

    nombre_empresa_seleccionada = "Todas las Empresas"
    if filtro_empresa == 'Samka':
        workers_queryset = workers_queryset.filter(empresa__nombre__icontains='Samka')
        nombre_empresa_seleccionada = "Samka SPA"
    elif filtro_empresa == 'Maquehue':
        workers_queryset = workers_queryset.filter(empresa__nombre__icontains='Maquehue')
        nombre_empresa_seleccionada = "Maquehue SPA"

    samka_activos = Trabajador.objects.filter(empresa__nombre__icontains='Samka', fecha_finiquito__isnull=True).count()
    samka_finiquitados = Trabajador.objects.filter(empresa__nombre__icontains='Samka', fecha_finiquito__isnull=False).count()
    maquehue_activos = Trabajador.objects.filter(empresa__nombre__icontains='Maquehue', fecha_finiquito__isnull=True).count()
    maquehue_finiquitados = Trabajador.objects.filter(empresa__nombre__icontains='Maquehue', fecha_finiquito__isnull=False).count()
    total_activos = samka_activos + maquehue_activos
    total_finiquitados = samka_finiquitados + maquehue_finiquitados

    activos_resumen = workers_queryset.values('empresa__nombre', 'cargo').annotate(
        total=Count('id'),
        activos=Count('id', filter=Q(fecha_finiquito__isnull=True)),
        finiquitados=Count('id', filter=Q(fecha_finiquito__isnull=False))
    ).order_by('empresa__nombre', '-activos')
    
    finiquitos = workers_queryset.filter(fecha_finiquito__isnull=False)\
                                 .annotate(mes=TruncMonth('fecha_finiquito'))\
                                 .values('mes')\
                                 .annotate(total=Sum('monto_finiquito'))\
                                 .order_by('mes')

    labels_grafico = [f.get('mes').strftime('%Y-%m') for f in finiquitos if f.get('mes')]
    data_grafico = [f.get('total') for f in finiquitos if f.get('mes')]

    lista_trabajadores = workers_queryset.order_by('empresa', 'nombre')

    if request.GET.get('modo_ajax') == 'true':
        html_tabla = render_to_string(
            'core/partials/tabla_trabajadores.html', 
            {'lista_trabajadores': lista_trabajadores},
            request=request
        )
        return JsonResponse({
            'html_tabla': html_tabla,
            'labels_grafico': labels_grafico,
            'data_grafico': data_grafico,
            'titulo_pagina': nombre_empresa_seleccionada
        })

    context = {
        'lista_trabajadores': lista_trabajadores,
        'activos_resumen': activos_resumen,
        'labels_grafico': labels_grafico,
        'data_grafico': data_grafico,
        'nombre_empresa': nombre_empresa_seleccionada,
        'samka_activos': samka_activos,
        'samka_finiquitados': samka_finiquitados,
        'maquehue_activos': maquehue_activos,
        'maquehue_finiquitados': maquehue_finiquitados,
        'total_activos': total_activos,
        'total_finiquitados': total_finiquitados,
    }
    return render(request, 'core/dashboard_rrhh.html', context)

@login_required
def importar_rrhh(request):
    """Importador Avanzado RRHH"""
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        archivo = request.FILES['archivo_excel']
        try:
            xls = pd.ExcelFile(archivo, engine='openpyxl')
            hojas = xls.sheet_names
            texto_hojas = "".join(str(h).upper() for h in hojas)
            
            empresa_archivo = None
            if "SAMKA" in texto_hojas and "MAQUEHUE" not in texto_hojas:
                empresa_archivo, _ = Empresa.objects.get_or_create(nombre="Samka SPA")
            elif "MAQUEHUE" in texto_hojas and "SAMKA" not in texto_hojas:
                empresa_archivo, _ = Empresa.objects.get_or_create(nombre="Maquehue SPA")
            
            trabajadores_batch = {} 
            
            for nombre_hoja in hojas:
                nombre_upper = str(nombre_hoja).upper()
                empresa_hoja = None
                if "SAMKA" in nombre_upper:
                    empresa_hoja, _ = Empresa.objects.get_or_create(nombre="Samka SPA")
                elif "MAQUEHUE" in nombre_upper:
                    empresa_hoja, _ = Empresa.objects.get_or_create(nombre="Maquehue SPA")
                
                if not empresa_hoja and ("FINIQUITADO" in nombre_upper or "PERSONAL" in nombre_upper):
                    empresa_hoja = empresa_archivo
                
                if not empresa_hoja: continue 

                df = pd.read_excel(archivo, sheet_name=nombre_hoja)
                df.columns = df.columns.str.strip().str.upper()

                if 'RUT' not in df.columns or 'NOMBRE' not in df.columns: continue

                es_hoja_finiquito = "FINIQUITADO" in nombre_upper or "PERSONAL" in nombre_upper

                for index, row in df.iterrows():
                    rut = str(row.get('RUT', '')).strip().upper()
                    if not rut or len(rut) < 3 or rut == 'NAN': continue

                    nombre = str(row.get('NOMBRE', '')).strip()
                    cargo_txt = str(row.get('CARGO', 'Operario')).strip()
                    if cargo_txt.upper() == 'NAN': cargo_txt = 'Operario'
                    
                    fecha_inicio = row.get('CONTRATO')
                    if pd.isnull(fecha_inicio): fecha_inicio = None
                    
                    fecha_fin = row.get('FINIQUITO')
                    if pd.isnull(fecha_fin) or isinstance(fecha_fin, (int, float)):
                         fecha_fin = None
                    
                    monto = 0
                    if 'FINIQUITO.1' in df.columns:
                        val_monto = row.get('FINIQUITO.1', 0)
                        if isinstance(val_monto, (int, float)) and not pd.isna(val_monto):
                            monto = val_monto
                    
                    estado_nuevo = 'ACTIVO'
                    if fecha_fin or es_hoja_finiquito:
                        estado_nuevo = 'FINIQUITADO'

                    if rut in trabajadores_batch:
                        previo = trabajadores_batch[rut]
                        if previo['estado'] == 'FINIQUITADO':
                            if not previo['fecha_contrato'] and fecha_inicio:
                                previo['fecha_contrato'] = fecha_inicio
                            if not previo['monto_finiquito'] and monto > 0:
                                previo['monto_finiquito'] = monto
                            if not previo['fecha_finiquito'] and fecha_fin:
                                previo['fecha_finiquito'] = fecha_fin
                        continue
                    
                    trabajadores_batch[rut] = {
                        'nombre': nombre,
                        'cargo_txt': cargo_txt,
                        'empresa': empresa_hoja,
                        'fecha_contrato': fecha_inicio,
                        'fecha_finiquito': fecha_fin,
                        'monto_finiquito': monto,
                        'estado': estado_nuevo
                    }

            creados = 0
            actualizados = 0
            
            with transaction.atomic():
                for rut, data in trabajadores_batch.items():
                    cargo_obj, _ = Cargo.objects.get_or_create(
                        nombre__iexact=data['cargo_txt'],
                        defaults={'nombre': data['cargo_txt']}
                    )
                    
                    defaults = {
                        'nombre': data['nombre'],
                        'cargo': cargo_obj,
                        'empresa': data['empresa'],
                        'fecha_contrato': data['fecha_contrato'],
                        'estado': data['estado']
                    }
                    
                    if data['fecha_finiquito']:
                        defaults['fecha_finiquito'] = data['fecha_finiquito']
                    elif data['estado'] == 'ACTIVO':
                        defaults['fecha_finiquito'] = None
                        
                    if data['monto_finiquito'] > 0:
                        defaults['monto_finiquito'] = data['monto_finiquito']

                    obj, created = Trabajador.objects.update_or_create(
                        rut=rut,
                        defaults=defaults
                    )
                    if created: creados += 1
                    else: actualizados += 1

            messages.success(request, f'Procesado correctamente: {creados} nuevos, {actualizados} actualizados.')
            return redirect('dashboard_rrhh')

        except Exception as e:
            print(f"ERROR IMPT: {e}")
            messages.error(request, f"Error al importar: {str(e)}")

    return render(request, 'core/importar_rrhh.html')

@login_required
def nuevo_trabajador(request):
    if request.method == 'POST':
        form = TrabajadorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Trabajador registrado.')
            return redirect('dashboard_rrhh')
    else:
        form = TrabajadorForm()
    return render(request, 'core/nuevo_trabajador.html', {'form': form})

@login_required
def editar_trabajador(request, id):
    trabajador = get_object_or_404(Trabajador, id=id)
    if request.method == 'POST':
        form = TrabajadorForm(request.POST, instance=trabajador)
        if form.is_valid():
            form.save()
            messages.success(request, f'Datos de {trabajador.nombre} actualizados.')
            return redirect('dashboard_rrhh')
    else:
        form = TrabajadorForm(instance=trabajador)
    return render(request, 'core/nuevo_trabajador.html', {'form': form, 'titulo': 'Editar Trabajador'})


# =========================================================
# 6. MÓDULO LOGÍSTICA / INVENTARIO
# =========================================================
@login_required
@user_passes_test(es_bodega)
def inventario_dashboard(request):
    # 1. Base Query (Traemos lotes con sus productos)
    lotes = Lote.objects.select_related('producto').all().order_by('fecha_vencimiento')

    # 2. Filtros
    # Búsqueda Texto
    q = request.GET.get('q')
    if q:
        lotes = lotes.filter(
            Q(producto__nombre__icontains=q) |
            Q(producto__codigo__icontains=q) |
            Q(numero_lote__icontains=q)
        )

    # Filtro Categoría
    categoria = request.GET.get('categoria')
    if categoria:
        lotes = lotes.filter(producto__categoria=categoria)

    # Filtro Estado (Semáforo)
    estado = request.GET.get('estado')
    hoy = datetime.date.today()
    if estado == 'vencido':
        lotes = lotes.filter(fecha_vencimiento__lt=hoy)
    elif estado == 'por_vencer':
        # Próximos 30 días
        limite = hoy + datetime.timedelta(days=30)
        lotes = lotes.filter(fecha_vencimiento__gte=hoy, fecha_vencimiento__lte=limite)
    elif estado == 'ok':
        lotes = lotes.filter(fecha_vencimiento__gt=hoy + datetime.timedelta(days=30))

    # 3. Datos para el Gráfico (Stock por Categoría)
    datos_grafico = lotes.values('producto__categoria').annotate(total_stock=Sum('cantidad')).order_by('-total_stock')
    
    labels_grafico = [d['producto__categoria'] for d in datos_grafico]
    data_grafico = [d['total_stock'] for d in datos_grafico]

    # 4. Paginación
    paginator = Paginator(lotes, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # 5. Respuesta AJAX
    if request.GET.get('modo_ajax'):
        html_tabla = render_to_string('core/partials/tabla_inventario.html', {'page_obj': page_obj}, request=request)
        html_paginacion = render_to_string('core/partials/paginacion.html', {'page_obj': page_obj}, request=request)
        
        return JsonResponse({
            'html_tabla': html_tabla,
            'html_paginacion': html_paginacion,
            'grafico_labels': labels_grafico,
            'grafico_data': data_grafico
        })

    # 6. Respuesta Normal
    categorias = Producto.objects.values_list('categoria', flat=True).distinct()

    context = {
        'page_obj': page_obj,
        'categorias': categorias,
        'labels_grafico': labels_grafico,
        'data_grafico': data_grafico,
        'cat_sel': categoria,
        'estado_sel': estado,
    }
    return render(request, 'core/inventario/dashboard.html', context)

@login_required
def ingresar_lote(request):
    if request.method == 'POST':
        form = LoteForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Lote ingresado correctamente.')
            return redirect('inventario_dashboard')
    else:
        form = LoteForm()
    
    return render(request, 'core/inventario/form_lote.html', {'form': form})


# =========================================================
# 7. MÓDULO USUARIOS Y LOGIN
# =========================================================
@login_required
def registro_usuario(request):
    if not request.user.is_superuser:
        messages.error(request, 'Acceso denegado.')
        return redirect('dashboard')

    if request.method == 'POST':
        form = RegistroUsuarioForm(request.POST)
        if form.is_valid():
            user = form.save()
            try:
                grupo = Group.objects.get(name='Digitadores')
                user.groups.add(grupo)
            except Group.DoesNotExist: pass
            messages.success(request, f'Usuario {user.username} creado.')
            return redirect('dashboard')
    else:
        form = RegistroUsuarioForm()
    return render(request, 'registration/registro.html', {'form': form})

@login_required
def perfil_usuario(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user) 
            messages.success(request, 'Contraseña actualizada.')
            return redirect('perfil_usuario')
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'core/perfil.html', {'form': form})

class CustomLoginView(LoginView):
    """Login para usuarios normales (Azul)"""
    template_name = 'core/login.html'
    redirect_authenticated_user = True 

    def form_valid(self, form):
        recuerdame = self.request.POST.get('recuerdame')
        if recuerdame:
            self.request.session.set_expiry(1209600)
        else:
            self.request.session.set_expiry(0)
        return super().form_valid(form)

class AdminLoginView(CustomLoginView):
    """Login para administradores (Oscuro)"""
    template_name = 'core/login_admin.html'
    redirect_authenticated_user = False


# =========================================================
# 8. MÓDULO INTELIGENCIA ARTIFICIAL
# =========================================================
@login_required
def api_entrenar_ia(request):
    exito, mensaje = entrenar_modelo()
    return JsonResponse({'status': 'ok' if exito else 'error', 'mensaje': mensaje})

@login_required
def api_predecir_categoria(request):
    texto = request.GET.get('texto', '')
    sugerencia = predecir_categoria(texto)
    return JsonResponse({'categoria': sugerencia if sugerencia else None})

@login_required
@user_passes_test(es_bodega)
def salida_stock(request):
    """Descuenta stock usando lógica FIFO"""
    if request.method == 'POST':
        form = SalidaStockForm(request.POST)
        if form.is_valid():
            producto = form.cleaned_data['producto']
            cantidad_solicitada = form.cleaned_data['cantidad']
            precio_total = form.cleaned_data['precio_total']

            stock_actual = producto.lote_set.aggregate(total=Sum('cantidad'))['total'] or 0
            
            if cantidad_solicitada > stock_actual:
                messages.error(request, f'Error: Stock insuficiente. Tienes {stock_actual}, intentas vender {cantidad_solicitada}.')
            else:
                try:
                    with transaction.atomic():
                        # A. Lógica FIFO
                        lotes = Lote.objects.filter(producto=producto).order_by('fecha_vencimiento')
                        pendiente = cantidad_solicitada
                        
                        for lote in lotes:
                            if pendiente <= 0: break
                            
                            if lote.cantidad <= pendiente:
                                pendiente -= lote.cantidad
                                lote.delete()
                            else:
                                lote.cantidad -= pendiente
                                lote.save()
                                pendiente = 0

                        # B. Lógica FINANCIERA
                        Ingreso.objects.create(
                            fecha=datetime.date.today(),
                            tipo_documento='VENTA', 
                            monto_transferencia=precio_total,
                            descripcion_movimiento=f"Venta de {cantidad_solicitada} x {producto.nombre}",
                            detalle="Generado automáticamente desde Inventario",
                            clasificacion=None, 
                            empresa=None 
                        )
                    
                    messages.success(request, f'¡Venta registrada! Stock descontado y ${precio_total} ingresados a caja.')
                    return redirect('inventario_dashboard')

                except Exception as e:
                    messages.error(request, f"Error al procesar la venta: {e}")

    else:
        form = SalidaStockForm()

    return render(request, 'core/inventario/form_salida.html', {'form': form})

@login_required
def enviar_alerta_vencimientos(request):
    """Revisa lotes por vencer y envía un correo"""
    hoy = datetime.date.today()
    fecha_limite = hoy + datetime.timedelta(days=30) 

    lotes_vencidos = Lote.objects.filter(fecha_vencimiento__lt=hoy)
    lotes_por_vencer = Lote.objects.filter(fecha_vencimiento__range=[hoy, fecha_limite])

    if not lotes_vencidos.exists() and not lotes_por_vencer.exists():
        messages.info(request, 'No hay productos en riesgo para reportar.')
        return redirect('inventario_dashboard')

    try:
        asunto = f"⚠️ ALERTA DE STOCK - {hoy.strftime('%d/%m/%Y')}"
        
        mensaje_html = render_to_string('core/emails/alerta_stock.html', {
            'lotes_vencidos': lotes_vencidos,
            'lotes_por_vencer': lotes_por_vencer,
            'usuario': request.user
        })

        send_mail(
            subject=asunto,
            message="",
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[request.user.email], 
            html_message=mensaje_html,
            fail_silently=False,
        )

        messages.success(request, f'Informe enviado correctamente a {request.user.email}')
    
    except Exception as e:
        messages.error(request, f'Error al enviar correo: {str(e)}')
        print(f"Error Email: {e}")

    return redirect('inventario_dashboard')

@login_required
def centro_datos(request):
    return render(request, 'core/exportar_datos.html')

@login_required
def exportar_finanzas_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="dataset_finanzas.csv"'

    writer = csv.writer(response)
    writer.writerow(['ID', 'Fecha', 'Año', 'Mes', 'Tipo', 'Empresa', 'Centro Costo', 'Clasificacion', 'Descripcion', 'Detalle', 'Monto'])

    movimientos = Ingreso.objects.select_related('empresa', 'centro_costo', 'clasificacion').all().order_by('-fecha')

    for mov in movimientos:
        writer.writerow([
            mov.id,
            mov.fecha,
            mov.fecha.year,
            mov.fecha.month,
            mov.tipo_documento,
            mov.empresa.nombre if mov.empresa else 'Sin Asignar',
            mov.centro_costo.nombre if mov.centro_costo else 'General',
            mov.clasificacion.nombre if mov.clasificacion else 'Sin Clasificar',
            mov.descripcion_movimiento,
            mov.detalle,
            mov.monto_transferencia
        ])

    return response

@login_required
def exportar_inventario_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="dataset_stock_actual.csv"'

    writer = csv.writer(response)
    writer.writerow(['SKU', 'Producto', 'Categoria', 'Nro Lote', 'Fecha Vencimiento', 'Dias para Vencer', 'Estado', 'Cantidad Stock'])

    lotes = Lote.objects.select_related('producto').all().order_by('fecha_vencimiento')
    hoy = datetime.date.today()

    for lote in lotes:
        dias = (lote.fecha_vencimiento - hoy).days
        estado = "VENCIDO" if dias < 0 else "POR VENCER" if dias <= 30 else "OK"

        writer.writerow([
            lote.producto.codigo,
            lote.producto.nombre,
            lote.producto.categoria,
            lote.numero_lote,
            lote.fecha_vencimiento,
            dias,
            estado,
            lote.cantidad
        ])

    return response

@login_required
def nuevo_ingreso(request):
    if request.method == 'POST':
        form = IngresoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Registro guardado correctamente.')
            return redirect('dashboard')
    else:
        form = IngresoForm()
    
    return render(request, 'core/nuevo_ingreso.html', {'form': form})

def exportar_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Financiero"

    headers = ['ID', 'Fecha', 'Tipo', 'Categoría', 'Descripción', 'Monto']
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    from .models import Movimiento 
    movimientos = Movimiento.objects.all().order_by('-fecha')

    for mov in movimientos:
        ws.append([
            mov.id,
            mov.fecha.strftime('%d/%m/%Y'),
            mov.tipo,
            str(mov.categoria),
            mov.descripcion,
            mov.monto
        ])

    dim_holder = {}
    for col in range(ws.min_column, ws.max_column + 1):
        dim_holder[col] = 0
        
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                dim_holder[cell.column] = max((dim_holder[cell.column], len(str(cell.value))))
    
    for col, width in dim_holder.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Reporte_Finanzas.xlsx"'
    
    wb.save(response)
    return response